import pandas as pd
import glob
import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import threading
import random

class EtutListesiProgrami:
    def __init__(self, root):
        self.root = root
        self.root.title("Etüt Listesi Oluşturucu")
        self.root.geometry("1000x800")
        self.root.resizable(True, True)
        
        # Modern Material Design renk paleti
        self.colors = {
            'bg': '#F3F4F6',           # Arka plan - çok açık gri
            'card': '#FFFFFF',          # Kartlar - tam beyaz
            'primary': '#3B82F6',       # Mavi - ana renk
            'primary_hover': '#2563EB',  # Mavi hover
            'success': '#10B981',       # Yeşil - başarı
            'success_hover': '#059669', # Yeşil hover
            'danger': '#EF4444',        # Kırmızı - tehlike
            'danger_hover': '#DC2626',  # Kırmızı hover
            'text_primary': '#111827',  # Ana metin - koyu gri
            'text_secondary': '#6B7280', # İkincil metin - orta gri
            'text_light': '#9CA3AF',    # Açık metin
            'border': '#E5E7EB',        # Kenarlık - açık gri
            'border_dashed': '#D1D5DB', # Kesikli çizgi - orta gri
            'accent': '#8B5CF6'         # Vurgu rengi
        }
        
        # Root arka plan rengi
        self.root.configure(bg=self.colors['bg'])
        
        # Seçilen dosyalar
        self.secili_dosyalar = []
        
        # Arayüzü oluştur
        self.arayuz_olustur()
        
    def arayuz_olustur(self):
        # Ana container - merkezi kart
        main_container = tk.Frame(self.root, bg=self.colors['bg'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)
        
        # Merkezi kart
        card = tk.Frame(main_container, bg=self.colors['card'], relief=tk.FLAT)
        card.pack(fill=tk.BOTH, expand=True)
        
        # İçerik padding
        content_frame = tk.Frame(card, bg=self.colors['card'], padx=40, pady=30)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Ana başlık - Modern font
        baslik_label = tk.Label(
            content_frame,
            text="Etüt Listesi Oluşturucu",
            font=("Segoe UI", 36, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text_primary']
        )
        baslik_label.pack(pady=(0, 30))
        
        # Dosya yönetimi bölümü - Drag & Drop ve Liste birlikte
        dosya_yonetim_frame = tk.Frame(content_frame, bg=self.colors['card'])
        dosya_yonetim_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Sol taraf: Drag & Drop alanı
        drop_container = tk.Frame(dosya_yonetim_frame, bg=self.colors['card'])
        drop_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        drop_label = tk.Label(
            drop_container,
            text="Dosya Ekleme",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text_primary']
        )
        drop_label.pack(anchor=tk.W, pady=(0, 8))
        
        drop_frame = tk.Frame(drop_container, bg=self.colors['card'])
        drop_frame.pack(fill=tk.BOTH, expand=True)
        
        self.drop_canvas = tk.Canvas(
            drop_frame,
            bg="#FAFBFC",
            highlightthickness=0,
            relief=tk.FLAT,
            height=180
        )
        self.drop_canvas.pack(fill=tk.BOTH, expand=True)
        
        # Drop alanı metni ve ikon
        drop_text_frame = tk.Frame(self.drop_canvas, bg="#FAFBFC")
        self.drop_text_frame_id = self.drop_canvas.create_window(0, 0, window=drop_text_frame, anchor="center", tags="drop_content")
        
        # İkon (basit metin ikon)
        icon_label = tk.Label(
            drop_text_frame,
            text="📁",
            font=("Segoe UI", 36),
            bg="#FAFBFC",
            fg=self.colors['text_light']
        )
        icon_label.pack()
        
        self.drop_text = tk.Label(
            drop_text_frame,
            text="Drag & Drop ile\ndosya seçin",
            font=("Segoe UI", 11),
            fg=self.colors['text_secondary'],
            bg="#FAFBFC"
        )
        self.drop_text.pack(pady=(8, 0))
        
        # Kesikli çizgili kenarlık çiz
        self.drop_canvas.bind("<Configure>", self.ciz_drop_alani)
        
        # Sağ taraf: Dosya listesi
        liste_container_frame = tk.Frame(dosya_yonetim_frame, bg=self.colors['card'])
        liste_container_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        liste_label = tk.Label(
            liste_container_frame,
            text="Eklenen Dosyalar",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text_primary']
        )
        liste_label.pack(anchor=tk.W, pady=(0, 8))
        
        liste_container = tk.Frame(liste_container_frame, bg=self.colors['card'])
        liste_container.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar
        scrollbar = tk.Scrollbar(liste_container, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Liste kutusu - Modern stil (çoklu seçim, yuvarlatılmış görünüm)
        listbox_frame = tk.Frame(liste_container, bg="#FAFBFC", relief=tk.FLAT)
        listbox_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        self.dosya_listbox = tk.Listbox(
            listbox_frame,
            yscrollcommand=scrollbar.set,
            font=("Segoe UI", 11),
            height=8,
            bg="#FFFFFF",
            fg=self.colors['text_primary'],
            selectbackground=self.colors['primary'],
            selectforeground="white",
            relief=tk.FLAT,
            bd=0,
            highlightthickness=0,
            selectmode=tk.EXTENDED  # Çoklu seçim
        )
        self.dosya_listbox.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
        scrollbar.config(command=self.dosya_listbox.yview)
        
        # Butonlar - Modern buton tasarımı (Drag & Drop altında)
        buton_frame = tk.Frame(drop_container, bg=self.colors['card'])
        buton_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Buton stili fonksiyonu (ultra modern, yuvarlatılmış köşeler efekti)
        def modern_buton(parent, text, command, bg_color, hover_color, fg_color="white", width=15):
            # Modern buton için frame (gölge efekti için)
            btn_frame = tk.Frame(parent, bg=parent.cget('bg'))
            btn = tk.Button(
                btn_frame,
                text=text,
                command=command,
                font=("Segoe UI", 11, "bold"),
                bg=bg_color,
                fg=fg_color,
                padx=24,
                pady=14,
                cursor="hand2",
                relief=tk.FLAT,
                bd=0,
                activebackground=hover_color,
                activeforeground=fg_color,
                width=width,
                borderwidth=0,
                highlightthickness=0
            )
            btn.pack(fill=tk.BOTH, expand=True)
            
            # Hover efekti (yumuşak geçiş + hafif büyüme)
            def on_enter(e):
                btn.config(bg=hover_color)
                btn_frame.config(bg=hover_color)
            def on_leave(e):
                btn.config(bg=bg_color)
                btn_frame.config(bg=parent.cget('bg'))
            btn.bind("<Enter>", on_enter)
            btn.bind("<Leave>", on_leave)
            btn_frame.bind("<Enter>", on_enter)
            btn_frame.bind("<Leave>", on_leave)
            return btn_frame
        
        dosya_ekle_btn = modern_buton(
            buton_frame,
            "➕ Dosya Ekle",
            self.dosya_ekle,
            self.colors['primary'],
            self.colors['primary_hover'],
            width=14
        )
        dosya_ekle_btn.pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        
        klasor_ekle_btn = modern_buton(
            buton_frame,
            "📂 Klasör",
            self.klasor_ekle,
            self.colors['success'],
            self.colors['success_hover'],
            width=14
        )
        klasor_ekle_btn.pack(side=tk.LEFT, padx=(0, 8), fill=tk.X, expand=True)
        
        dosya_sil_btn = modern_buton(
            buton_frame,
            "🗑️ Kaldır",
            self.dosya_sil,
            self.colors['danger'],
            self.colors['danger_hover'],
            width=14
        )
        dosya_sil_btn.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # İşlem butonu - Büyük ve dikkat çekici (liste altında)
        islem_frame = tk.Frame(content_frame, bg=self.colors['card'])
        islem_frame.pack(fill=tk.X, pady=(20, 0))
        
        # Ana işlem butonu - Ultra modern
        islem_btn_frame = tk.Frame(islem_frame, bg=self.colors['card'])
        islem_btn_frame.pack()
        
        self.islem_btn = tk.Button(
            islem_btn_frame,
            text="🚀 Etüt Listesini Oluştur",
            command=self.islem_baslat,
            font=("Segoe UI", 18, "bold"),
            bg=self.colors['success'],
            fg="white",
            padx=60,
            pady=20,
            cursor="hand2",
            relief=tk.FLAT,
            bd=0,
            activebackground=self.colors['success_hover'],
            activeforeground="white",
            borderwidth=0,
            highlightthickness=0
        )
        self.islem_btn.pack()
        
        # Hover efekti için işlem butonu
        def on_enter_btn(e):
            self.islem_btn.config(bg=self.colors['success_hover'])
        def on_leave_btn(e):
            self.islem_btn.config(bg=self.colors['success'])
        self.islem_btn.bind("<Enter>", on_enter_btn)
        self.islem_btn.bind("<Leave>", on_leave_btn)
        
        # İlerleme çubuğu - Modern stil
        progress_frame = tk.Frame(islem_frame, bg=self.colors['card'])
        progress_frame.pack(fill=tk.X, pady=(20, 0))
        
        self.progress = ttk.Progressbar(
            progress_frame,
            mode='indeterminate',
            length=500,
            style="Modern.Horizontal.TProgressbar"
        )
        self.progress.pack()
        
        # Durum etiketi - Modern font
        self.durum_label = tk.Label(
            progress_frame,
            text="Hazır",
            font=("Segoe UI", 11),
            fg=self.colors['text_secondary'],
            bg=self.colors['card']
        )
        self.durum_label.pack(pady=(10, 0))
        
        # Bilgi kutusu - Durum label'ının altında, dönen bilgiler (animasyonlu)
        bilgi_kutusu = tk.Frame(
            progress_frame,
            bg="#ECFDF5",
            relief=tk.FLAT,
            padx=20,
            pady=16
        )
        bilgi_kutusu.pack(fill=tk.X, pady=(15, 0))
        
        # Animasyon için Canvas
        bilgi_canvas = tk.Canvas(
            bilgi_kutusu,
            bg="#ECFDF5",
            highlightthickness=0,
            height=50
        )
        bilgi_canvas.pack(fill=tk.BOTH, expand=True)
        
        # İki label (çıkış ve giriş için)
        self.bilgi_text_old = tk.Label(
            bilgi_canvas,
            text="",
            font=("Segoe UI", 11),
            fg="#065F46",
            bg="#ECFDF5",
            wraplength=900,
            justify=tk.CENTER
        )
        self.bilgi_text_new = tk.Label(
            bilgi_canvas,
            text="",
            font=("Segoe UI", 11),
            fg="#065F46",
            bg="#ECFDF5",
            wraplength=900,
            justify=tk.CENTER
        )
        
        # Dönen bilgiler listesi
        self.bilgi_mesajlari = [
            "💡 Bu uygulama sınav analiz dosyalarınızı işleyerek otomatik etüt listesi oluşturur",
            "📋 Dosya Formatı: Excel (.xlsx, .xls) veya CSV formatında sınav analiz dosyaları kullanılmalıdır",
            "📊 Puanlama: Sınıf başarısı %35 ve altındaysa TÜM SINIF, üstündeyse bireysel etüt uygulanır",
            "⏰ Etüt Süreleri: Bireysel Etüt → 20 dakika (5 soru) | Sınıf Etütü → 40 dakika (5 soru)",
            "👥 Grup Dağılımı: Etüt grupları maksimum 4 kişi olacak şekilde dengeli dağıtılır",
            "✅ Çoğunluk Kuralı: Bir kazanımda öğrencilerin %50'den fazlası bireysel etüt alıyorsa sınıf etütü yapılır",
            "📁 Toplu İşlem: Birden fazla dosyayı aynı anda seçebilir veya klasör ekleyebilirsiniz",
            "🎯 Otomatik Analiz: Program dosyalarınızı analiz ederek hangi öğrencilerin hangi konulardan etüde kalacağını belirler"
        ]
        self.bilgi_index = 0
        self.bilgi_animasyon_aktif = False
        self.bilgi_canvas = bilgi_canvas
        self.bilgi_guncelle()
        
        # Modern progressbar stili
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Modern.Horizontal.TProgressbar",
                       background=self.colors['success'],
                       troughcolor="#E5E7EB",
                       borderwidth=0,
                       lightcolor=self.colors['success'],
                       darkcolor=self.colors['success'])
    
    def bilgi_guncelle(self):
        """Bilgi mesajlarını 5 saniyede bir döndür (animasyonlu geçiş)"""
        if not hasattr(self, 'bilgi_mesajlari') or not hasattr(self, 'bilgi_canvas'):
            return
        
        if self.bilgi_animasyon_aktif:
            return
        
        # Canvas boyutunu al
        self.bilgi_canvas.update_idletasks()
        canvas_width = self.bilgi_canvas.winfo_width()
        canvas_height = self.bilgi_canvas.winfo_height()
        
        if canvas_width < 10:
            # Canvas henüz boyutlanmamış, tekrar dene
            self.root.after(100, self.bilgi_guncelle)
            return
        
        # Yeni mesaj
        yeni_mesaj = self.bilgi_mesajlari[self.bilgi_index]
        self.bilgi_index = (self.bilgi_index + 1) % len(self.bilgi_mesajlari)
        
        # Eski label varsa animasyonla çıkar
        if hasattr(self, 'bilgi_text_old_id') and self.bilgi_text_old_id:
            try:
                # Eski label'ın var olup olmadığını kontrol et
                self.bilgi_canvas.coords(self.bilgi_text_old_id)  # Hata verirse yok demektir
            except:
                # İlk mesaj - animasyon yok
                self.bilgi_text_new.config(text=yeni_mesaj)
                self.bilgi_text_new_id = self.bilgi_canvas.create_window(
                    canvas_width / 2, canvas_height / 2,
                    window=self.bilgi_text_new, anchor="center"
                )
                self.bilgi_text_old_id = self.bilgi_text_new_id
                self.root.after(5000, self.bilgi_guncelle)
                return
            
            self.bilgi_animasyon_aktif = True
            eski_x = canvas_width / 2
            yeni_x = -canvas_width / 2
            
            # Eski label'ı sola kaydır
            def animasyon_cikis(step=0):
                if step <= 15:
                    x = eski_x - (eski_x - yeni_x) * (step / 15)
                    try:
                        self.bilgi_canvas.coords(self.bilgi_text_old_id, x, canvas_height / 2)
                    except:
                        pass
                    self.root.after(15, lambda: animasyon_cikis(step + 1))
                else:
                    # Eski label'ı sil
                    try:
                        self.bilgi_canvas.delete(self.bilgi_text_old_id)
                    except:
                        pass
                    
                    # Yeni label'ı sağdan getir
                    self.bilgi_text_new.config(text=yeni_mesaj)
                    self.bilgi_text_new_id = self.bilgi_canvas.create_window(
                        canvas_width * 1.5, canvas_height / 2,
                        window=self.bilgi_text_new, anchor="center"
                    )
                    
                    def animasyon_giris(step=0):
                        if step <= 15:
                            x = canvas_width * 1.5 - (canvas_width * 1.5 - canvas_width / 2) * (step / 15)
                            try:
                                self.bilgi_canvas.coords(self.bilgi_text_new_id, x, canvas_height / 2)
                            except:
                                pass
                            self.root.after(15, lambda: animasyon_giris(step + 1))
                        else:
                            # Yeni label'ı eski yap
                            self.bilgi_text_old_id = self.bilgi_text_new_id
                            self.bilgi_text_old, self.bilgi_text_new = self.bilgi_text_new, self.bilgi_text_old
                            self.bilgi_animasyon_aktif = False
                            self.root.after(5000, self.bilgi_guncelle)  # 5 saniye sonra tekrar
                    
                    animasyon_giris()
            
            animasyon_cikis()
        else:
            # İlk mesaj - animasyon yok
            self.bilgi_text_new.config(text=yeni_mesaj)
            self.bilgi_text_new_id = self.bilgi_canvas.create_window(
                canvas_width / 2, canvas_height / 2,
                window=self.bilgi_text_new, anchor="center"
            )
            self.bilgi_text_old_id = self.bilgi_text_new_id
            self.root.after(5000, self.bilgi_guncelle)  # 5 saniye sonra tekrar
    
    def ciz_drop_alani(self, event=None):
        """Drag & Drop alanı için kesikli çizgili kenarlık çizer"""
        self.drop_canvas.delete("border")
        width = self.drop_canvas.winfo_width()
        height = self.drop_canvas.winfo_height()
        
        if width > 1 and height > 1:
            # Canvas içeriğini merkeze al
            self.drop_canvas.coords(self.drop_text_frame_id, width/2, height/2)
            
            # Kesikli çizgi efekti (dash pattern)
            dash = (8, 4)
            self.drop_canvas.create_rectangle(
                10, 10, width-10, height-10,
                outline=self.colors['border_dashed'],
                width=2,
                dash=dash,
                tags="border"
            )
        
    def dosya_ekle(self):
        dosyalar = filedialog.askopenfilenames(
            title="Sınav Analiz Dosyalarını Seçin",
            filetypes=[
                ("Excel/CSV Dosyaları", "*.xlsx *.xls *.csv"),
                ("Excel Dosyaları", "*.xlsx *.xls"),
                ("CSV Dosyaları", "*.csv"),
                ("Tüm Dosyalar", "*.*")
            ]
        )
        
        for dosya in dosyalar:
            if dosya not in self.secili_dosyalar:
                self.secili_dosyalar.append(dosya)
                self.dosya_listbox.insert(tk.END, f"📄 {os.path.basename(dosya)}")
        
        self.durum_guncelle(f"✅ {len(self.secili_dosyalar)} dosya seçildi")
        
    def klasor_ekle(self):
        klasor = filedialog.askdirectory(title="Klasör Seçin")
        if klasor:
            dosyalar = []
            for ext in ['*.xlsx', '*.xls', '*.csv']:
                dosyalar.extend(glob.glob(os.path.join(klasor, ext)))
            
            eklenen = 0
            for dosya in dosyalar:
                if dosya not in self.secili_dosyalar:
                    self.secili_dosyalar.append(dosya)
                    self.dosya_listbox.insert(tk.END, f"📄 {os.path.basename(dosya)}")
                    eklenen += 1
            
            if eklenen > 0:
                self.durum_guncelle(f"✅ {eklenen} dosya eklendi")
            else:
                messagebox.showinfo("Bilgi", "Klasörde uygun dosya bulunamadı.")
        
    def dosya_sil(self):
        secili_indeksler = self.dosya_listbox.curselection()
        if not secili_indeksler:
            messagebox.showwarning("Uyarı", "Lütfen silmek için en az bir dosya seçin.")
            return
        
        # Tersten sil (çoklu seçim desteği)
        for indeks in reversed(secili_indeksler):
            self.dosya_listbox.delete(indeks)
            del self.secili_dosyalar[indeks]
        
        self.durum_guncelle(f"📋 {len(self.secili_dosyalar)} dosya kaldı")
        
    def durum_guncelle(self, mesaj):
        self.durum_label.config(text=mesaj)
        self.root.update()
        
    def islem_baslat(self):
        if not self.secili_dosyalar:
            messagebox.showwarning("Uyarı", "Lütfen en az bir dosya seçin!")
            return
        
        # Butonu devre dışı bırak
        self.islem_btn.config(state=tk.DISABLED, text="⏳ İşleniyor...")
        self.progress.start()
        self.durum_guncelle("İşlem başlatılıyor...")
        
        # Arka planda çalıştır
        thread = threading.Thread(target=self.islem_yap)
        thread.daemon = True
        thread.start()
        
    def islem_yap(self):
        try:
            tum_data = []
            basarili_dosya_sayisi = 0
            
            for i, dosya_yolu in enumerate(self.secili_dosyalar):
                dosya_adi = os.path.basename(dosya_yolu)
                self.durum_guncelle(f"📊 İşleniyor: {dosya_adi} ({i+1}/{len(self.secili_dosyalar)})")
                
                try:
                    sonuc = self.data_kent_analiz(dosya_yolu)
                    if sonuc:
                        tum_data.extend(sonuc)
                        basarili_dosya_sayisi += 1
                        
                        # Dosyalar arası boş satır ekle (son dosya değilse)
                        if i < len(self.secili_dosyalar) - 1:
                            tum_data.append({
                                "Dosya": "",
                                "Soru": "",
                                "Kazanım": "",
                                "Etüt Grubu": "",
                                "Öğrenciler": "",
                                "Sebep": "",
                                "Etüt Süresi": "",
                                "Soru Sayısı": "",
                                "Etüt Tipi": ""
                            })
                except Exception as e:
                    print(f"Hata: {dosya_yolu} - {str(e)}")
            
            if tum_data:
                # Çıktı dosyasını kaydet
                cikti_dosyasi = filedialog.asksaveasfilename(
                    title="Etüt Listesini Kaydet",
                    defaultextension=".xlsx",
                    filetypes=[("Excel Dosyası", "*.xlsx"), ("Tüm Dosyalar", "*.*")]
                )
                
                if cikti_dosyasi:
                    df_sonuc = pd.DataFrame(tum_data)
                    
                    # Excel'e yazarken PROFESYONEL ve CANLI formatlama
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    
                    with pd.ExcelWriter(cikti_dosyasi, engine='openpyxl') as writer:
                        df_sonuc.to_excel(writer, index=False, sheet_name='Etüt Planı')
                        
                        worksheet = writer.sheets['Etüt Planı']
                        
                        # MODERN VE CANLI TASARIM
                        # 1. Font: Segoe UI (daha büyük ve okunabilir)
                        segoe_font = Font(name='Segoe UI', size=11, color='2C3E50')
                        segoe_font_bold = Font(name='Segoe UI', bold=True, size=11, color='2C3E50')
                        segoe_font_header = Font(name='Segoe UI', bold=True, color='FFFFFF', size=12)
                        
                        # 2. Başlık satırı: Canlı mavi gradient efekti (#2563EB - daha canlı)
                        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
                        
                        # 3. Kenarlıklar: Daha belirgin (orta kalınlık, koyu gri)
                        border_style = Border(
                            left=Side(style='medium', color='D1D5DB'),
                            right=Side(style='medium', color='D1D5DB'),
                            top=Side(style='medium', color='D1D5DB'),
                            bottom=Side(style='medium', color='D1D5DB')
                        )
                        border_style_thick = Border(
                            left=Side(style='thick', color='9CA3AF'),
                            right=Side(style='thick', color='9CA3AF'),
                            top=Side(style='thick', color='9CA3AF'),
                            bottom=Side(style='thick', color='9CA3AF')
                        )
                        
                        # 4. Hizalama tanımlamaları
                        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        # Başlık satırını formatla (daha canlı)
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = segoe_font_header
                            cell.alignment = center_align
                            cell.border = border_style_thick
                        
                        # Sütun genişliklerini ayarla (daha geniş)
                        column_widths = {
                            'A': 38,  # Dosya
                            'B': 10,  # Soru
                            'C': 60,  # Kazanım
                            'D': 14,  # Etüt Grubu
                            'E': 55,  # Öğrenciler
                            'F': 85,  # Sebep (detaylı)
                            'G': 18,  # Etüt Süresi
                            'H': 14,  # Soru Sayısı
                            'I': 20   # Etüt Tipi
                        }
                        
                        for col, width in column_widths.items():
                            worksheet.column_dimensions[col].width = width
                        
                        # Satırları formatla (daha canlı renkler)
                        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                            # Boş satır kontrolü
                            is_empty = all(cell.value is None or str(cell.value).strip() == '' for cell in row)
                            
                            # Etüt Tipi ve Öğrenciler sütunlarını kontrol et (koşullu biçimlendirme için)
                            etut_tipi = None
                            ogrenciler = None
                            if len(row) > 8:
                                etut_tipi_cell = row[8]  # Etüt Tipi sütunu (I sütunu, index 8)
                                if etut_tipi_cell.value:
                                    etut_tipi = str(etut_tipi_cell.value).strip()
                            if len(row) > 4:
                                ogrenciler_cell = row[4]  # Öğrenciler sütunu (E sütunu, index 4)
                                if ogrenciler_cell.value:
                                    ogrenciler = str(ogrenciler_cell.value).strip()
                            
                            # Koşullu biçimlendirme: Sınıf Etütü veya TÜM SINIF -> Canlı kırmızı/pembe (#FECACA - daha canlı)
                            is_sinif_etutu = (etut_tipi == "Sınıf Etütü") or (ogrenciler == "TÜM SINIF")
                            # Koşullu biçimlendirme: Bireysel Etüt -> Canlı mavi (#DBEAFE - daha canlı)
                            is_bireysel_etut = (etut_tipi == "Bireysel Etüt")
                            
                            for cell in row:
                                cell.border = border_style
                                
                                if is_empty:
                                    # Boş satırları farklı renkle işaretle (daha yumuşak)
                                    cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                                    cell.font = Font(name='Segoe UI', size=10, color='9CA3AF')
                                else:
                                    # Koşullu biçimlendirme uygula (daha canlı renkler)
                                    if is_sinif_etutu:
                                        cell.fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
                                        cell.font = Font(name='Segoe UI', size=11, color='991B1B', bold=True)
                                    elif is_bireysel_etut:
                                        cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                                        cell.font = Font(name='Segoe UI', size=11, color='1E40AF')
                                    else:
                                        cell.font = segoe_font
                                    
                                    # Sütun bazlı hizalama ve formatlama
                                    if cell.column == 1:  # Dosya sütunu - Yatay ortalı
                                        cell.alignment = center_align
                                        cell.font = Font(name='Segoe UI', bold=True, size=11, color='1F2937')
                                    elif cell.column == 2:  # Soru sütunu - Yatay ortalı
                                        cell.alignment = center_align
                                        cell.font = Font(name='Segoe UI', bold=True, size=12, color='2563EB')
                                    elif cell.column == 3:  # Kazanım sütunu - Sola dayalı, Wrap Text
                                        cell.alignment = left_align
                                        if not is_sinif_etutu and not is_bireysel_etut:
                                            cell.font = segoe_font
                                    elif cell.column == 4:  # Etüt Grubu - Yatay ortalı
                                        cell.alignment = center_align
                                        cell.font = Font(name='Segoe UI', bold=True, size=12, color='059669')
                                    elif cell.column == 5:  # Öğrenciler - Sola dayalı, Wrap Text
                                        cell.alignment = left_align
                                        if not is_sinif_etutu and not is_bireysel_etut:
                                            cell.font = segoe_font
                                    elif cell.column == 6:  # Sebep - Sola dayalı, Wrap Text
                                        cell.alignment = left_align
                                        if not is_sinif_etutu and not is_bireysel_etut:
                                            cell.font = Font(name='Segoe UI', size=10, color='4B5563')
                                    elif cell.column == 7:  # Etüt Süresi - Yatay ortalı
                                        cell.alignment = center_align
                                        if not is_sinif_etutu and not is_bireysel_etut:
                                            cell.font = Font(name='Segoe UI', size=10, color='059669', bold=True)
                                    elif cell.column == 8:  # Soru Sayısı - Yatay ortalı
                                        cell.alignment = center_align
                                        if not is_sinif_etutu and not is_bireysel_etut:
                                            cell.font = Font(name='Segoe UI', size=10, color='059669', bold=True)
                                    elif cell.column == 9:  # Etüt Tipi - Yatay ortalı
                                        cell.alignment = center_align
                                        if not is_sinif_etutu and not is_bireysel_etut:
                                            cell.font = Font(name='Segoe UI', bold=True, size=11, color='7C3AED')
                        
                        # Satır yüksekliklerini ayarla (daha yüksek - daha okunabilir)
                        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                            if row[0].row == 1:
                                worksheet.row_dimensions[row[0].row].height = 35  # Başlık daha yüksek
                            else:
                                worksheet.row_dimensions[row[0].row].height = 28  # Normal satırlar daha yüksek
                    
                    self.progress.stop()
                    self.islem_btn.config(state=tk.NORMAL, text="🚀 Etüt Listesini Oluştur")
                    self.durum_guncelle("✅ İşlem tamamlandı!")
                    
                    # Toplam etüt grubu sayısını hesapla
                    toplam_grup = len([x for x in tum_data if x.get("Etüt Grubu") and str(x.get("Etüt Grubu")) != ""])
                    
                    messagebox.showinfo(
                        "Başarılı",
                        f"✅ İşlem tamamlandı!\n\n"
                        f"📊 Toplam {toplam_grup} etüt grubu oluşturuldu\n"
                        f"📁 {basarili_dosya_sayisi} dosya başarıyla işlendi\n"
                        f"💾 Dosya kaydedildi: {os.path.basename(cikti_dosyasi)}"
                    )
                else:
                    self.progress.stop()
                    self.islem_btn.config(state=tk.NORMAL, text="🚀 Etüt Listesini Oluştur")
                    self.durum_guncelle("İptal edildi")
            else:
                self.progress.stop()
                self.islem_btn.config(state=tk.NORMAL, text="🚀 Etüt Listesini Oluştur")
                self.durum_guncelle("Etüt kaydı bulunamadı")
                messagebox.showwarning(
                    "Uyarı",
                    "Hiçbir dosyadan etüt kaydı çıkarılamadı.\nDosya formatlarını kontrol edin."
                )
                
        except Exception as e:
            self.progress.stop()
            self.islem_btn.config(state=tk.NORMAL, text="🚀 Etüt Listesini Oluştur")
            self.durum_guncelle("❌ Hata oluştu!")
            messagebox.showerror("Hata", f"Bir hata oluştu:\n{str(e)}")
    
    def data_kent_analiz(self, dosya_yolu):
        """
        Data Kent formatındaki sınav analiz dosyasını okuyup etüt listesi çıkarır.
        Geliştirilmiş: Kazanım ve puan okuma daha esnek ve doğru.
        """
        try:
            df_raw = pd.read_csv(dosya_yolu, header=None, engine='python', encoding='utf-8-sig')
        except:
            try:
                df_raw = pd.read_excel(dosya_yolu, header=None, engine='openpyxl')
            except:
                return []

        etut_listesi = []
        dosya_adi = os.path.basename(dosya_yolu)

        # 1. ADIM: "ADI VE SOYADI" başlık satırını bul
        baslik_index = -1
        for i, row in df_raw.iterrows():
            satir = " ".join([str(x) for x in row.values if pd.notna(x)])
            if "ADI VE SOYADI" in satir.upper():
                baslik_index = i
                break
                
        if baslik_index == -1:
            return []

        # Öğrenci tablosunu oluştur
        df_students = df_raw.iloc[baslik_index+1:].copy()
        df_students.columns = df_raw.iloc[baslik_index]
        
        # "ADI VE SOYADI" sütununu bul
        adi_soyadi_col = None
        for col in df_students.columns:
            if 'ADI VE SOYADI' in str(col).upper():
                adi_soyadi_col = col
                break
        
        # Sıra numarası sütununu bul
        sira_no_col = None
        for col in df_students.columns:
            col_str = str(col).upper()
            if 'SIRA' in col_str and 'NO' in col_str:
                sira_no_col = col
                break
        
        # Sadece gerçek öğrenci satırlarını al
        if sira_no_col:
            sira_sutun = df_students[sira_no_col]
            mask = pd.to_numeric(sira_sutun, errors='coerce').notnull()
        else:
            ilk_sutun = df_students.iloc[:, 0]
            mask = pd.to_numeric(ilk_sutun, errors='coerce').notnull()
        
        if adi_soyadi_col:
            adi_mask = df_students[adi_soyadi_col].notna()
            try:
                kazanim_mask = ~df_students[adi_soyadi_col].astype(str).str.contains('Kazanım', case=False, na=False)
                mask = mask & adi_mask & kazanim_mask
            except:
                mask = mask & adi_mask
        
        df_students = df_students[mask].copy()
        df_students = df_students.dropna(how='all')
        
        if df_students.empty:
            return []

        # 2. ADIM: "SORULARA GÖRE BAŞARI (%)" satırını bul
        sinif_basari_yuzdeleri = {}
        
        for i, row in df_raw.iterrows():
            satir = " ".join([str(x) for x in row.values if pd.notna(x)])
            if "SORULARA GÖRE BAŞARI" in satir.upper() or "BAŞARI (%)" in satir.upper():
                baslik_row = df_raw.iloc[baslik_index]
                
                for col_idx, val in enumerate(row.values):
                    if col_idx < len(baslik_row):
                        baslik_val = baslik_row.iloc[col_idx]
                        baslik_str = str(baslik_val).strip() if pd.notna(baslik_val) else ""
                        baslik_clean = baslik_str.replace('.', '').replace(',', '')
                        if baslik_clean.isdigit():
                            soru_no = str(int(float(baslik_str))) if '.' in baslik_str else baslik_str
                            if pd.notna(val):
                                try:
                                    if isinstance(val, (int, float)):
                                        yuzde_val = float(val)
                                    else:
                                        val_str = str(val).strip().replace('%', '').replace(',', '.')
                                        yuzde_match = re.search(r'(\d+\.?\d*)', val_str)
                                        if yuzde_match:
                                            yuzde_val = float(yuzde_match.group(1))
                                        else:
                                            continue
                                    if 0 <= yuzde_val <= 100:
                                        sinif_basari_yuzdeleri[soru_no] = yuzde_val
                                except:
                                    continue
                break

        # 3. ADIM: "Soruların ilgili olduğu konular, kazanımlar..." başlığını bul ve altındaki satırları işle
        kazanim_map = {}
        soru_max_puan = {}
        
        # Başlık satırını bul: "Soruların ilgili olduğu konular, kazanımlar veya alt öğrenme alanları"
        kazanim_baslik_index = None
        puan_sutun_index = None
        
        for i, row in df_raw.iterrows():
            satir = " ".join([str(x) for x in row.values if pd.notna(x)])
            satir_upper = satir.upper()
            
            # Başlık satırını bul
            if "SORULAR" in satir_upper and ("KAZANIM" in satir_upper or "KONU" in satir_upper or "ÖĞRENME ALANI" in satir_upper):
                kazanim_baslik_index = i
                
                # "Puan" sütununu bul
                for idx, val in enumerate(row.values):
                    if pd.notna(val):
                        val_str = str(val).upper()
                        if "PUAN" in val_str:
                            puan_sutun_index = idx
                            break
                break
        
        # Başlık bulunduysa, altındaki satırları işle
        if kazanim_baslik_index is not None:
            # Başlık satırındaki sütun yapısını anla
            baslik_row = df_raw.iloc[kazanim_baslik_index]
            
            # Başlık satırındaki soru numarası sütununu bul (genelde ilk sütun veya başlık satırındaki sayısal sütunlar)
            soru_no_sutun_index = None
            for idx, val in enumerate(baslik_row.values):
                if pd.notna(val):
                    val_str = str(val).strip()
                    # Başlık satırında sayısal değer varsa, bu soru numarası sütunu olabilir
                    if val_str.replace('.', '').isdigit():
                        soru_no_sutun_index = idx
                        break
            
            # Eğer bulunamadıysa, ilk sütunu varsay
            if soru_no_sutun_index is None:
                soru_no_sutun_index = 0
            
            # Kazanım metni sütununu bul (başlık satırında "Soruların ilgili olduğu konular, kazanımlar..." yazan sütun)
            kazanim_metin_sutun_index = None
            for idx, val in enumerate(baslik_row.values):
                if pd.notna(val):
                    val_str = str(val).upper()
                    # "SORULAR" ve ("KONU" veya "KAZANIM" veya "ÖĞRENME") içeren sütun
                    if "SORULAR" in val_str and ("KONU" in val_str or "KAZANIM" in val_str or "ÖĞRENME" in val_str):
                        kazanim_metin_sutun_index = idx
                        break
                    # Veya sadece "KONU", "KAZANIM", "ÖĞRENME" içeren sütun (başlık satırında)
                    elif ("KONU" in val_str or "KAZANIM" in val_str or "ÖĞRENME" in val_str) and "PUAN" not in val_str:
                        if kazanim_metin_sutun_index is None:  # İlk bulduğunu al
                            kazanim_metin_sutun_index = idx
            
            # Eğer bulunamadıysa, soru numarası sütununun yanındaki sütunu dene
            if kazanim_metin_sutun_index is None:
                # Soru numarası sütununun yanındaki 3 sütunu kontrol et
                for offset in [1, 2, 3]:
                    test_idx = soru_no_sutun_index + offset
                    if test_idx < len(baslik_row.values):
                        test_val = baslik_row.iloc[test_idx]
                        if pd.notna(test_val):
                            test_str = str(test_val).upper()
                            # "PUAN" içermiyorsa ve boş değilse
                            if "PUAN" not in test_str and len(test_str) > 3:
                                kazanim_metin_sutun_index = test_idx
                                break
            
            # Başlık satırının altındaki satırları işle
            for i in range(kazanim_baslik_index + 1, len(df_raw)):
                row = df_raw.iloc[i]
                
                # Boş satır kontrolü - eğer satır tamamen boşsa veya sadece NaN varsa dur
                if row.isna().all():
                    continue
                
                try:
                    soru_no = None
                    kazanim_metni = None
                    max_puan = None
                    
                    # 1. SORU NUMARASINI BUL
                    if soru_no_sutun_index < len(row.values):
                        soru_val = row.iloc[soru_no_sutun_index]
                        if pd.notna(soru_val):
                            soru_str = str(soru_val).strip()
                            # Sayısal değer mi kontrol et
                            if soru_str.replace('.', '').isdigit():
                                soru_no = str(int(float(soru_str))) if '.' in soru_str else soru_str
                                # Mantıklı bir soru numarası mı (1-100)
                                if not (1 <= int(soru_no) <= 100):
                                    soru_no = None
                    
                    # Eğer hala bulunamadıysa, ilk sütundaki sayısal değere bak
                    if not soru_no and len(row.values) > 0:
                        first_val = row.iloc[0]
                        if pd.notna(first_val):
                            first_str = str(first_val).strip()
                            if first_str.replace('.', '').isdigit():
                                sayi = int(float(first_str)) if '.' in first_str else int(first_str)
                                if 1 <= sayi <= 100:
                                    soru_no = str(sayi)
                    
                    # 2. KAZANIM METNİNİ BUL
                    if kazanim_metin_sutun_index < len(row.values):
                        kazanim_val = row.iloc[kazanim_metin_sutun_index]
                        if pd.notna(kazanim_val):
                            kazanim_metni = str(kazanim_val).strip()
                            # Eğer çok kısaysa, sonraki sütunlara bak
                            if len(kazanim_metni) < 5:
                                # Sonraki 3 sütuna bak
                                for next_idx in range(kazanim_metin_sutun_index + 1, min(kazanim_metin_sutun_index + 4, len(row.values))):
                                    next_val = row.iloc[next_idx]
                                    if pd.notna(next_val):
                                        next_str = str(next_val).strip()
                                        # Sayısal değilse ve yeterince uzunsa
                                        if not next_str.replace('.', '').isdigit() and len(next_str) > 5:
                                            kazanim_metni = next_str
                                            break
                    
                    # 3. PUANI BUL
                    if puan_sutun_index is not None and puan_sutun_index < len(row.values):
                        puan_val = row.iloc[puan_sutun_index]
                        if pd.notna(puan_val):
                            try:
                                if isinstance(puan_val, (int, float)):
                                    max_puan = int(puan_val)
                                else:
                                    puan_str = str(puan_val).strip()
                                    if puan_str.replace('.', '').isdigit():
                                        max_puan = int(float(puan_str)) if '.' in puan_str else int(puan_str)
                            except:
                                pass
                    
                    # Puan bulunamadıysa, satırdaki diğer sayısal değerlere bak
                    if max_puan is None:
                        for idx, val in enumerate(row.values):
                            if pd.notna(val):
                                val_str = str(val).strip()
                                if val_str.replace('.', '').isdigit():
                                    sayi = int(float(val_str)) if '.' in val_str else int(val_str)
                                    # Soru numarası değilse ve mantıklı bir puan değeriyse
                                    if soru_no and str(sayi) == soru_no:
                                        continue
                                    if 5 <= sayi <= 100:
                                        # Kazanım metninin sağındaki sayılar puan olma ihtimali daha yüksek
                                        if kazanim_metin_sutun_index is not None and idx > kazanim_metin_sutun_index:
                                            max_puan = sayi
                                            break
                    
                    # Son çare: Öğrenci notlarından tahmin et
                    if max_puan is None and soru_no:
                        soru_col = None
                        for col in df_students.columns:
                            col_str = str(col).strip().split('.')[0]
                            if col_str == soru_no:
                                soru_col = col
                                break
                        
                        if soru_col:
                            notlar = pd.to_numeric(df_students[soru_col], errors='coerce').dropna()
                            if len(notlar) > 0:
                                max_not = notlar.max()
                                # Olası puan değerlerine yuvarla
                                for puan in [5, 10, 15, 20, 25, 30, 40, 50, 100]:
                                    if max_not <= puan:
                                        max_puan = puan
                                        break
                    
                    # Varsayılan: 25
                    if max_puan is None:
                        max_puan = 25
                    
                    # Kaydet: Soru numarası ve kazanım metni varsa
                    if soru_no and kazanim_metni and len(kazanim_metni) > 3:
                        kazanim_map[soru_no] = kazanim_metni
                        soru_max_puan[soru_no] = max_puan
                        
                except Exception as e:
                    continue

        # 4. ADIM: Soru sütunlarını bul
        soru_sutunlari = []
        for col in df_students.columns:
            col_str = str(col).strip()
            col_clean = col_str.replace('.', '').replace(',', '')
            if col_clean.isdigit():
                soru_sutunlari.append(col)

        if not soru_sutunlari:
            return []

        # 5. ADIM: Kuralları uygula
        for soru_col in soru_sutunlari:
            soru_no = str(soru_col).strip().split('.')[0]
            
            konu = kazanim_map.get(soru_no, f"{soru_no}. Soru")
            max_p = soru_max_puan.get(soru_no, 25)
            sinif_yuzde = sinif_basari_yuzdeleri.get(soru_no)
            
            # Eğer sınıf başarı yüzdesi bulunamadıysa, öğrenci notlarından hesapla
            if sinif_yuzde is None:
                notlar = pd.to_numeric(df_students[soru_col], errors='coerce').dropna()
                if len(notlar) > 0:
                    sinif_ort = notlar.mean()
                    sinif_yuzde = (sinif_ort / max_p) * 100 if max_p > 0 else 0
                else:
                    sinif_yuzde = 0

            # KURAL 1: Sınıf başarısı %35 ve altındaysa TÜM SINIF
            if sinif_yuzde <= 35:
                # %0 kontrolü - hata uyarısı
                sebep_metni = f"Sınıf Başarısı: %{sinif_yuzde:.1f} (≤%35)"
                if sinif_yuzde == 0:
                    sebep_metni += " ⚠️ UYARI: Bu kazanımda Excel dosyasında hata olabilir (Sınıf Başarısı %0)"
                
                etut_listesi.append({
                    "Dosya": dosya_adi,
                    "Soru": soru_no,
                    "Kazanım": konu,
                    "Etüt Grubu": 1,
                    "Öğrenciler": "TÜM SINIF",
                    "Sebep": sebep_metni,
                    "Etüt Süresi": "40 dakika",
                    "Soru Sayısı": "5 soru",
                    "Etüt Tipi": "Sınıf Etütü"
                })
                continue

            # KURAL 2: Bireysel kontrol
            limit = max_p * 0.5
            ogrenciler_detayli = []  # (isim, puan) tuple listesi
            sinif_mevcudu = len(df_students)  # Toplam öğrenci sayısı
            
            for _, ogrenci in df_students.iterrows():
                try:
                    puan = pd.to_numeric(ogrenci[soru_col], errors='coerce')
                    if pd.isna(puan):
                        continue
                    
                    if puan <= limit:
                        ogrenci_adi = None
                        if adi_soyadi_col:
                            ogrenci_adi_val = ogrenci.get(adi_soyadi_col)
                            if pd.notna(ogrenci_adi_val):
                                ogrenci_adi = str(ogrenci_adi_val).strip()
                                if 'Kazanım' in ogrenci_adi or 'KAZANIM' in ogrenci_adi:
                                    ogrenci_adi = None
                        
                        if not ogrenci_adi:
                            for col_name in df_students.columns:
                                col_str = str(col_name).upper()
                                if 'ADI' in col_str and 'SOYADI' in col_str:
                                    ogrenci_adi_val = ogrenci.get(col_name)
                                    if pd.notna(ogrenci_adi_val):
                                        ogrenci_adi_temp = str(ogrenci_adi_val).strip()
                                        if ogrenci_adi_temp and 'Kazanım' not in ogrenci_adi_temp:
                                            ogrenci_adi = ogrenci_adi_temp
                                            break
                        
                        if ogrenci_adi and ogrenci_adi.lower() != 'nan' and ogrenci_adi != '':
                            ogrenciler_detayli.append((ogrenci_adi, puan))
                except:
                    continue
            
            # KURAL 3: Çoğunluk kontrolü - Eğer bireysel etüt alan öğrenci sayısı sınıf mevcudunun %50'sinden fazlaysa, sınıf etütü yap
            if ogrenciler_detayli and sinif_mevcudu > 0:
                bireysel_etut_yuzdesi = (len(ogrenciler_detayli) / sinif_mevcudu) * 100
                if bireysel_etut_yuzdesi > 50:
                    # Sınıf etütü yap
                    sebep_metni = f"Sınıf Başarısı: %{sinif_yuzde:.1f} (>%35), ancak öğrencilerin %{bireysel_etut_yuzdesi:.1f}'i bireysel etüt alıyor (Çoğunluk Kuralı)"
                    if sinif_yuzde == 0:
                        sebep_metni += " ⚠️ UYARI: Bu kazanımda Excel dosyasında hata olabilir (Sınıf Başarısı %0)"
                    
                    etut_listesi.append({
                        "Dosya": dosya_adi,
                        "Soru": soru_no,
                        "Kazanım": konu,
                        "Etüt Grubu": 1,
                        "Öğrenciler": "TÜM SINIF",
                        "Sebep": sebep_metni,
                        "Etüt Süresi": "40 dakika",
                        "Soru Sayısı": "5 soru",
                        "Etüt Tipi": "Sınıf Etütü"
                    })
                    continue
            
            # Öğrencileri etüt gruplarına dengeli böl (maksimum 4 kişi, dengeli dağıtım)
            if ogrenciler_detayli:
                # Öğrenci listesini random karıştır (alfabetik sıra yerine, her seferinde farklı gruplar)
                random.shuffle(ogrenciler_detayli)
                
                toplam_ogrenci = len(ogrenciler_detayli)
                etut_grup_no = 1
                
                # Dengeli grup dağılımı algoritması
                def dengeli_grup_dagit(toplam):
                    """Dengeli grup dağılımı: 5→3+2, 6→3+3, 7→4+3, 8→4+4, 9→3+3+3, 10→4+3+3, 11→4+4+3, 12→4+4+4 vb."""
                    if toplam <= 4:
                        return [toplam]
                    
                    gruplar = []
                    kalan = toplam
                    
                    while kalan > 0:
                        if kalan <= 4:
                            gruplar.append(kalan)
                            break
                        elif kalan == 5:
                            gruplar.extend([3, 2])
                            break
                        elif kalan == 6:
                            gruplar.extend([3, 3])
                            break
                        elif kalan == 7:
                            gruplar.extend([4, 3])
                            break
                        elif kalan == 8:
                            gruplar.extend([4, 4])
                            break
                        elif kalan == 9:
                            gruplar.extend([3, 3, 3])
                            break
                        elif kalan == 10:
                            gruplar.extend([4, 3, 3])
                            break
                        elif kalan == 11:
                            gruplar.extend([4, 4, 3])
                            break
                        elif kalan == 12:
                            gruplar.extend([4, 4, 4])
                            break
                        else:
                            # 13 ve üzeri için: 4'lük gruplar oluştur, kalanı dengeli dağıt
                            if kalan % 4 == 0:
                                # Tam 4'lük gruplar
                                gruplar.extend([4] * (kalan // 4))
                                break
                            elif kalan % 4 == 1:
                                # Son grup 5 olacak, önceki gruplardan birini 3'e düşür
                                gruplar.extend([4] * ((kalan // 4) - 1))
                                gruplar.extend([3, 2])
                                break
                            elif kalan % 4 == 2:
                                # Son grup 6 olacak, önceki gruplardan birini 3'e düşür
                                gruplar.extend([4] * ((kalan // 4) - 1))
                                gruplar.extend([3, 3])
                                break
                            else:  # kalan % 4 == 3
                                # Son grup 7 olacak
                                gruplar.extend([4] * (kalan // 4))
                                gruplar.append(3)
                                break
                    
                    return gruplar
                
                grup_boyutlari = dengeli_grup_dagit(toplam_ogrenci)
                ogrenci_index = 0
                
                for grup_boyutu in grup_boyutlari:
                    grup_ogrenciler = ogrenciler_detayli[ogrenci_index:ogrenci_index + grup_boyutu]
                    ogrenci_index += grup_boyutu
                    
                    # Öğrenci isimlerini ve puanlarını formatla
                    ogrenci_isimleri = [og[0] for og in grup_ogrenciler]
                    ogrenci_puanlari = [f"{og[0]} ({og[1]:.1f}/{max_p})" for og in grup_ogrenciler]
                    
                    ogrenciler_str = ", ".join(ogrenci_isimleri)
                    sebep_detay = " | ".join(ogrenci_puanlari)
                    
                    etut_listesi.append({
                        "Dosya": dosya_adi,
                        "Soru": soru_no,
                        "Kazanım": konu,
                        "Etüt Grubu": etut_grup_no,
                        "Öğrenciler": ogrenciler_str,
                        "Sebep": sebep_detay,
                        "Etüt Süresi": "20 dakika",
                        "Soru Sayısı": "5 soru",
                        "Etüt Tipi": "Bireysel Etüt"
                    })
                    etut_grup_no += 1

        return etut_listesi


def main():
    root = tk.Tk()
    app = EtutListesiProgrami(root)
    root.mainloop()


if __name__ == "__main__":
    main()
